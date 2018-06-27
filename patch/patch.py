# coding: utf-8
from openpyxl.reader import excel

from .excel import _load_workbook

setattr(excel, '_load_workbook', _load_workbook)
