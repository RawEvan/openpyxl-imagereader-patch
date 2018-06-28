# coding: utf-8

from openpyxl.reader.excel import load_workbook
from openpyxl.reader.excel import _validate_archive, _find_workbook_part
from openpyxl.packaging.workbook import WorkbookParser
from .drawings import read_drawings, get_drawings_file
from openpyxl.xml.constants import ARC_CONTENT_TYPES
from openpyxl.xml.functions import fromstring
from openpyxl.packaging.manifest import Manifest


def load_workbook_img(
        filename,
        read_only=False,
        keep_vba=False,
        guess_types=False,
        data_only=False,
):
    archive = _validate_archive(filename)

    src = archive.read(ARC_CONTENT_TYPES)
    root = fromstring(src)
    package = Manifest.from_tree(root)

    wb = load_workbook(filename, read_only, keep_vba, guess_types, data_only)
    valid_files = archive.namelist()
    wb_sheet_names = wb.sheetnames
    wb_part = _find_workbook_part(package)
    parser = WorkbookParser(archive, wb_part.PartName[1:])
    parser.parse()
    for sheet, rel in parser.find_sheets():
        sheet_name = sheet.name
        ws = wb[sheet_name]
        worksheet_path = rel.target
        if not worksheet_path in valid_files:
            continue
        if not sheet_name in wb_sheet_names:
            continue
        drawings_file = get_drawings_file(worksheet_path, archive, valid_files)
        if drawings_file is not None:
            ws = wb[sheet_name]
            read_drawings(ws, drawings_file, archive, valid_files)
    return wb
